# app/exports.py
from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from flask import Response, send_file

# XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BytesLike = Union[bytes, bytearray, memoryview]
TextLike = Union[str, bytes]

_SAFE_NAME_RE = re.compile(r"[^a-zA-Z0-9._-]+")
_SAFE_SLUG_RE = re.compile(r"[^a-zA-Z0-9_-]+")


# ----------------------------
# Paths (central)
# ----------------------------
BASE_DIR: Path = Path(__file__).resolve().parents[1]  # repo root
CONFIG_DIR: Path = BASE_DIR / "config"
EXPORTS_DIR: Path = BASE_DIR / "exports"
CERTS_DIR: Path = BASE_DIR / "certs"
EXPORT_CONFIG_PATH: Path = CONFIG_DIR / "exports.json"


# ----------------------------
# Small utils
# ----------------------------
def deep_merge(defaults: dict, overrides: dict) -> dict:
    """Recursively merge overrides into defaults."""
    if not isinstance(defaults, dict):
        return overrides
    if not isinstance(overrides, dict):
        return defaults

    out = dict(defaults)
    for k, v in overrides.items():
        if k in out and isinstance(out[k], dict) and isinstance(v, dict):
            out[k] = deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def ensure_exports_dir() -> None:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def slugify_filename(name: str, default: str = "export") -> str:
    base = (Path(name).stem or "").strip() or default
    base = _SAFE_SLUG_RE.sub("_", base)
    base = base.strip("_") or default
    return base[:120]


def safe_filename(name: str, default: str = "download") -> str:
    """Makes a safe filename for Content-Disposition."""
    name = (name or "").strip() or default
    name = name.replace("\\", "_").replace("/", "_")
    name = _SAFE_NAME_RE.sub("_", name)
    return name[:180]


def to_bytes(data: TextLike, encoding: str = "utf-8") -> bytes:
    if isinstance(data, (bytes, bytearray, memoryview)):
        return bytes(data)
    return str(data).encode(encoding, errors="replace")


# ----------------------------
# Download helpers
# ----------------------------
def send_text_download(filename: str, text: str, mimetype: str = "text/plain; charset=utf-8") -> Response:
    filename = safe_filename(filename)
    bio = io.BytesIO(to_bytes(text))
    bio.seek(0)
    return send_file(bio, mimetype=mimetype, as_attachment=True, download_name=filename)


def send_bytes_download(filename: str, data: BytesLike, mimetype: str = "application/octet-stream") -> Response:
    filename = safe_filename(filename)
    bio = io.BytesIO(bytes(data))
    bio.seek(0)
    return send_file(bio, mimetype=mimetype, as_attachment=True, download_name=filename)


def csv_bytes(rows: Sequence[dict], fieldnames: Optional[List[str]] = None, delimiter: str = ";") -> bytes:
    """
    Creates CSV bytes (UTF-8 with BOM for Excel friendliness).
    - delimiter defaults to ';' (BE/Excel)
    """
    if not rows:
        fieldnames = fieldnames or []
    else:
        if fieldnames is None:
            seen: List[str] = []
            for k in rows[0].keys():
                seen.append(k)
            for r in rows[1:]:
                for k in r.keys():
                    if k not in seen:
                        seen.append(k)
            fieldnames = seen

    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=fieldnames, delimiter=delimiter, lineterminator="\n")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k, "") for k in fieldnames})

    # BOM for Excel
    return ("\ufeff" + out.getvalue()).encode("utf-8", errors="replace")


def send_csv_download(
    filename: str,
    rows: Sequence[dict],
    fieldnames: Optional[List[str]] = None,
    delimiter: str = ";",
) -> Response:
    filename = safe_filename(filename)
    if not filename.lower().endswith(".csv"):
        filename += ".csv"
    data = csv_bytes(rows, fieldnames=fieldnames, delimiter=delimiter)
    return send_bytes_download(filename, data, mimetype="text/csv; charset=utf-8")


def zip_from_files(files: Iterable[Tuple[str, BytesLike]]) -> bytes:
    """files: iterable of (path_inside_zip, bytes)"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, content in files:
            name = name.replace("\\", "/").lstrip("/")
            z.writestr(name, bytes(content))
    return buf.getvalue()


def zip_from_folder(
    folder: Path,
    include_globs: Optional[List[str]] = None,
    exclude_globs: Optional[List[str]] = None,
) -> bytes:
    """
    Zips a folder recursively.
    include_globs: patterns like ["*.txt", "*.csv"] (if None => include all)
    exclude_globs: patterns like ["*.log", "__pycache__/*"]
    """
    folder = folder.resolve()
    include_globs = include_globs or []
    exclude_globs = exclude_globs or []

    def is_excluded(rel: str) -> bool:
        rel_norm = rel.replace("\\", "/")
        for pat in exclude_globs:
            if Path(rel_norm).match(pat) or Path(rel_norm).as_posix().startswith(pat.rstrip("*")):
                return True
        return False

    def is_included(p: Path) -> bool:
        if not include_globs:
            return True
        rel = p.relative_to(folder).as_posix()
        for pat in include_globs:
            if Path(rel).match(pat) or p.name.lower() == pat.lower():
                return True
        return False

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in folder.rglob("*"):
            if p.is_dir():
                continue
            rel = p.relative_to(folder).as_posix()
            if is_excluded(rel):
                continue
            if not is_included(p):
                continue
            z.write(p, arcname=rel)
    return buf.getvalue()


def send_zip_download(filename: str, zipped: bytes) -> Response:
    filename = safe_filename(filename)
    if not filename.lower().endswith(".zip"):
        filename += ".zip"
    return send_bytes_download(filename, zipped, mimetype="application/zip")


# ----------------------------
# Export styles (config/exports.json)
# ----------------------------
def default_export_styles(settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> Dict[str, Any]:
    settings = settings or {}
    branding = branding or {}

    app_title = (branding.get("app_title") or "Centraal Portaal").strip()
    copyright_text = (branding.get("copyright") or "© CyNiT 2024 - 2026").strip()

    # neutral defaults (can be overridden in config/exports.json)
    title_color = "#0a84ff"
    body_bg = "#ffffff"
    body_fg = "#111111"

    return {
        "meta": {
            "app_title": app_title,
            "copyright": copyright_text,
        },
        "xlsx": {
            "sheet": {"default_bg": "#FFFFFF"},
            "title": {"font_size": 16, "bold": True, "italic": False, "color": "#000000"},
            "field_col": {"font_size": 12, "bold": True, "italic": True, "color": "#000000"},
            "value_col": {"font_size": 12, "bold": False, "italic": False, "color": "#000000"},
        },
        "html": {
            "body": {"bg": body_bg, "fg": body_fg},
            "title": {"color": title_color, "font_size_px": 18, "bold": True},
            "table": {
                "border_color": "#000000",
                "border_width_px": 1,
                "field_col": {"font_size_px": 13, "bold": True, "italic": True},
                "value_col": {"font_size_px": 12, "bold": False, "italic": False},
            },
        },
        "md": {
            "doc_title": f"{app_title} Export",
            "title_prefix": "# ",
            "section_prefix": "## ",
            "bold_field_names": True,
            "table_header": "| Field | Value |",
            "table_sep": "| --- | --- |",
            "footer": copyright_text,
        },
        "csv": {
            "delimiter": ";",
            "header": ["Section", "Field", "Value"],
        },
    }


def load_export_styles(settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> Dict[str, Any]:
    defaults = default_export_styles(settings, branding)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

    if not EXPORT_CONFIG_PATH.exists():
        EXPORT_CONFIG_PATH.write_text(json.dumps(defaults, indent=2, ensure_ascii=False), encoding="utf-8")
        return defaults

    try:
        current = json.loads(EXPORT_CONFIG_PATH.read_text(encoding="utf-8"))
        merged = deep_merge(defaults, current)
    except Exception:
        merged = defaults

    # normalize file on disk
    try:
        EXPORT_CONFIG_PATH.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

    return merged


# ----------------------------
# Export builders (central)
# ----------------------------
def _export_title(branding: Dict[str, Any] | None, suffix: str = "Export") -> str:
    branding = branding or {}
    app_title = (branding.get("app_title") or "Centraal Portaal").strip()
    return f"{app_title} {suffix}".strip()


def build_csv_text(info: Dict[str, Any], settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> str:
    styles = load_export_styles(settings or {}, branding or {})
    delim = (styles.get("csv", {}) or {}).get("delimiter", ";") or ";"
    header = (styles.get("csv", {}) or {}).get("header") or ["Section", "Field", "Value"]

    lines = [delim.join(header)]

    checks = info.get("checks")
    if isinstance(checks, list):
        for c in checks:
            val = f'{c.get("status","")} - {c.get("message","")}'
            lines.append(delim.join(["Checks", str(c.get("name", "")), str(val).replace(delim, ",")]))

    for sec_key, sec_name in [("subject", "Subject"), ("issuer", "Issuer"), ("properties", "Properties")]:
        section = info.get(sec_key)
        if not isinstance(section, dict):
            continue
        for k, v in section.items():
            lines.append(delim.join([sec_name, str(k), str(v).replace(delim, ",")]))

    return "\n".join(lines) + "\n"


def build_html_export(info: Dict[str, Any], settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> str:
    settings = settings or {}
    branding = branding or {}
    styles = load_export_styles(settings, branding)

    html_cfg = styles["html"]
    body = html_cfg["body"]
    title_cfg = html_cfg["title"]
    table_cfg = html_cfg["table"]

    page_title = _export_title(branding, "Certificate Export")
    meta_app = (styles.get("meta", {}) or {}).get("app_title") or (branding.get("app_title") or "Centraal Portaal")
    meta_copyright = (styles.get("meta", {}) or {}).get("copyright") or (branding.get("copyright") or "© CyNiT 2024 - 2026")

    def table_block(title: str, mapping: Optional[Dict[str, Any]], is_issuer: bool = False) -> str:
        if is_issuer and mapping is None:
            return f"<h2>{title}</h2><p>CSR heeft geen issuer.</p>"
        if mapping is None:
            return ""

        rows = []
        for k, v in mapping.items():
            rows.append(
                f"<tr>"
                f"<td style='font-weight:bold;font-style:italic;font-size:{table_cfg['field_col']['font_size_px']}px'>{k}</td>"
                f"<td style='font-size:{table_cfg['value_col']['font_size_px']}px'>{v}</td>"
                f"</tr>"
            )
        return f"<h2>{title}</h2><table>{''.join(rows)}</table>"

    filename = info.get("filename", "")
    kind = info.get("kind") or info.get("type") or ""

    return f"""<!DOCTYPE html>
<html lang="nl">
<head>
<meta charset="utf-8">
<title>{page_title}</title>
<style>
body {{
  background:{body['bg']};
  color:{body['fg']};
  font-family:Arial, sans-serif;
  margin: 18px;
}}
table {{
  border-collapse:collapse;
  min-width:520px;
}}
td {{
  border:{table_cfg['border_width_px']}px solid {table_cfg['border_color']};
  padding:6px 10px;
  vertical-align: top;
}}
h1,h2 {{
  color:{title_cfg['color']};
  font-weight:bold;
}}
h1 {{
  font-size:{title_cfg['font_size_px']}px;
  margin: 0 0 8px 0;
}}
.small {{
  opacity: 0.75;
  font-size: 0.95rem;
}}
</style>
</head>
<body>

<h1>{meta_app} – Certificate Export</h1>
<div class="small">{meta_copyright}</div>

<p><strong>Bestand:</strong> {filename}</p>
<p><strong>Type:</strong> {kind}</p>

{table_block("Subject", info.get("subject"))}
{table_block("Issuer", info.get("issuer"), is_issuer=True)}
{table_block("Properties", info.get("properties"))}

</body>
</html>
"""


def build_markdown_export(info: Dict[str, Any], settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> str:
    settings = settings or {}
    branding = branding or {}
    styles = load_export_styles(settings, branding)

    md_cfg = styles["md"]
    title = (md_cfg.get("title_prefix") or "# ") + (md_cfg.get("doc_title") or _export_title(branding, "Certificate Export"))
    section = md_cfg.get("section_prefix") or "## "

    table_header = md_cfg.get("table_header", "| Field | Value |")
    table_sep = md_cfg.get("table_sep", "| --- | --- |")
    footer = md_cfg.get("footer", "")

    filename = info.get("filename", "")
    kind = info.get("kind") or info.get("type") or ""

    def md_table(title_txt: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> str:
        if issuer and mapping is None:
            return f"{section}{title_txt}\n\nCSR heeft geen issuer.\n"
        if mapping is None:
            return ""

        lines = [f"{section}{title_txt}", "", table_header, table_sep]
        for k, v in mapping.items():
            field = f"**{k}**" if md_cfg.get("bold_field_names", True) else str(k)
            lines.append(f"| {field} | {v} |")
        lines.append("")
        return "\n".join(lines)

    md_parts: List[str] = [
        title,
        "",
        f"**Bestand:** `{filename}`",
        f"**Type:** {kind}",
        "",
        md_table("Subject", info.get("subject")),
        md_table("Issuer", info.get("issuer"), issuer=True),
        md_table("Properties", info.get("properties")),
    ]

    checks = info.get("checks")
    if isinstance(checks, list) and checks:
        md_parts.append(f"{section}Checks\n")
        md_parts.append(table_header)
        md_parts.append(table_sep)
        for c in checks:
            k = c.get("name", "")
            v = f'{c.get("status","")} - {c.get("message","")}'
            field = f"**{k}**" if md_cfg.get("bold_field_names", True) else str(k)
            md_parts.append(f"| {field} | {v} |")
        md_parts.append("")

    if footer:
        md_parts.append(f"---\n{footer}\n")

    return "\n".join([p for p in md_parts if p is not None]).strip() + "\n"


def build_xlsx_export(info: Dict[str, Any], settings: Dict[str, Any] | None = None, branding: Dict[str, Any] | None = None) -> bytes:
    settings = settings or {}
    branding = branding or {}
    styles = load_export_styles(settings, branding)
    cfg = styles["xlsx"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    default_bg = (cfg["sheet"]["default_bg"] or "#FFFFFF").lstrip("#")
    fill = PatternFill(start_color=default_bg, end_color=default_bg, fill_type="solid")

    title_font = Font(
        size=int(cfg["title"]["font_size"]),
        bold=bool(cfg["title"]["bold"]),
        italic=bool(cfg["title"]["italic"]),
        color=str(cfg["title"]["color"]).lstrip("#"),
    )
    field_font = Font(
        size=int(cfg["field_col"]["font_size"]),
        bold=bool(cfg["field_col"]["bold"]),
        italic=bool(cfg["field_col"]["italic"]),
        color=str(cfg["field_col"]["color"]).lstrip("#"),
    )
    value_font = Font(
        size=int(cfg["value_col"]["font_size"]),
        bold=bool(cfg["value_col"]["bold"]),
        italic=bool(cfg["value_col"]["italic"]),
        color=str(cfg["value_col"]["color"]).lstrip("#"),
    )

    row = 1

    ws["A1"] = _export_title(branding, "Certificate Export")
    ws["A1"].font = title_font
    ws["A1"].fill = fill
    row += 2

    def write_row(key: str, value: Any) -> None:
        nonlocal row
        ws[f"A{row}"] = key
        ws[f"B{row}"] = "" if value is None else str(value)
        ws[f"A{row}"].font = field_font
        ws[f"B{row}"].font = value_font
        ws[f"A{row}"].fill = fill
        ws[f"B{row}"].fill = fill
        row += 1

    write_row("Bestand", info.get("filename", ""))
    write_row("Type", info.get("kind") or info.get("type") or "")
    row += 1

    def write_section(title: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> None:
        nonlocal row
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = title_font
        ws[f"A{row}"].fill = fill
        row += 1

        if issuer and mapping is None:
            write_row("Issuer", "CSR heeft geen issuer.")
            row += 1
            return

        for k, v in (mapping or {}).items():
            write_row(str(k), v)
        row += 1

    write_section("Subject", info.get("subject"))
    write_section("Issuer", info.get("issuer"), issuer=True)
    write_section("Properties", info.get("properties"))

    checks = info.get("checks")
    if isinstance(checks, list) and checks:
        ws[f"A{row}"] = "Checks"
        ws[f"A{row}"].font = title_font
        ws[f"A{row}"].fill = fill
        row += 1
        for c in checks:
            write_row(str(c.get("name", "")), f'{c.get("status","")} - {c.get("message","")}')
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 70

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem.getvalue()


def build_zip_bytes(
    info: Dict[str, Any],
    settings: Dict[str, Any] | None = None,
    branding: Dict[str, Any] | None = None,
    formats: Optional[List[str]] = None,
) -> bytes:
    """formats example: ["json","csv","html","md","xlsx"]"""
    settings = settings or {}
    branding = branding or {}
    formats = [f.lower().strip() for f in (formats or ["json", "csv", "html", "md", "xlsx"]) if f]

    base = slugify_filename(info.get("filename", "export"), default="export")
    mem = io.BytesIO()

    with zipfile.ZipFile(mem, "w", compression=zipfile.ZIP_DEFLATED) as z:
        if "json" in formats:
            z.writestr(f"{base}.json", json.dumps(info, indent=2, ensure_ascii=False))

        if "csv" in formats:
            z.writestr(f"{base}.csv", build_csv_text(info, settings, branding))

        if "html" in formats:
            z.writestr(f"{base}.html", build_html_export(info, settings, branding))

        if "md" in formats:
            z.writestr(f"{base}.md", build_markdown_export(info, settings, branding))

        if "xlsx" in formats:
            z.writestr(f"{base}.xlsx", build_xlsx_export(info, settings, branding))

    mem.seek(0)
    return mem.getvalue()


def read_file_bytes(path: Path, max_mb: int = 25) -> bytes:
    """Safe-ish file read with size cap."""
    path = path.resolve()
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))
    size = path.stat().st_size
    if size > max_mb * 1024 * 1024:
        raise ValueError(f"File too large ({size} bytes), limit is {max_mb} MB")
    return path.read_bytes()
