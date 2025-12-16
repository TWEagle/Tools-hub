#!/usr/bin/env python3
"""
dcb_org_export.py

CyNiT Tools module:
- Leest per omgeving config uit config/dcbaas_api.json
- Leest per omgeving auth data uit token_file (access_token.txt als JSON)
- Kan een client_assertion JWT genereren op basis van JWK (zoals JWT2JWK)
- Kan via client_credentials + client_assertion een nieuw access_token
  opvragen bij authenticatie(-ti).vlaanderen.be /op/v1/token
- Roept /certificate/search aan per organisatie-code
- Bouwt een Excel met alle toepassingen + certificaten

Integratie in ctools.py:
    import dcb_org_export
    dcb_org_export.register_web_routes(app, SETTINGS, TOOLS)
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
from io import BytesIO
import os
import json
import time
import datetime as dt

import requests
import jwt
from jwt.algorithms import RSAAlgorithm
from flask import Flask, request, render_template_string, send_file

import cynit_theme
import cynit_layout
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------- DEBUG ----------------

DEBUG = True  # zet op False als je minder console output wil


def log_debug(msg: str) -> None:
    if DEBUG:
        print(f"[DCBAAS ORG EXPORT] {msg}")


# Basis paden
BASE_DIR = cynit_theme.BASE_DIR
CONFIG_DIR = cynit_theme.CONFIG_DIR
DCBAAS_API_CFG = CONFIG_DIR / "dcbaas_api.json"


# ------------------------------------------------------------
#  Config / environment
# ------------------------------------------------------------

@dataclass
class EnvConfig:
    name: str
    label: str
    external_api_base: str
    access_token: Optional[str]
    token_file: Optional[str]
    auth_audience: Optional[str]
    token_url: Optional[str]
    scope: Optional[str]


def _write_skeleton_dcbaas_api() -> None:
    """
    Maak een skeleton dcbaas_api.json aan als hij nog niet bestaat.
    Dit is enkel om iets te hebben dat je via /config-editor kunt invullen.
    """
    skeleton = {
        "default_env": "TI",
        "environments": {
            "DEV": {
                "label": "DCBaaS DEV",
                "external_api_base": "",
                "access_token": "",
                "token_file": "",
                "auth_audience": "",
                "token_url": "",
                "scope": ""
            },
            "TI": {
                "label": "DCBaaS TI",
                "external_api_base": "",
                "access_token": "",
                "token_file": "",
                "auth_audience": "",
                "token_url": "",
                "scope": ""
            },
            "PROD": {
                "label": "DCBaaS PROD",
                "external_api_base": "",
                "access_token": "",
                "token_file": "",
                "auth_audience": "",
                "token_url": "",
                "scope": ""
            }
        }
    }
    DCBAAS_API_CFG.parent.mkdir(parents=True, exist_ok=True)
    DCBAAS_API_CFG.write_text(json.dumps(skeleton, indent=2), encoding="utf-8")
    log_debug(f"Skeleton dcbaas_api.json aangemaakt op {DCBAAS_API_CFG}")


def load_env_configs_from_dcbaas_api() -> Tuple[Dict[str, EnvConfig], Optional[str]]:
    """
    Leest config/dcbaas_api.json en bouwt EnvConfig per environment.
    """
    if not DCBAAS_API_CFG.exists():
        log_debug("dcbaas_api.json bestond niet – skeleton wordt aangemaakt.")
        _write_skeleton_dcbaas_api()

    try:
        raw = json.loads(DCBAAS_API_CFG.read_text(encoding="utf-8"))
        log_debug(f"dcbaas_api.json geladen: keys={list(raw.keys())}")
    except Exception as exc:
        log_debug(f"FOUT bij lezen dcbaas_api.json: {exc} – skeleton opnieuw schrijven.")
        _write_skeleton_dcbaas_api()
        raw = json.loads(DCBAAS_API_CFG.read_text(encoding="utf-8"))

    envs_raw = raw.get("environments", {})
    envs: Dict[str, EnvConfig] = {}

    for env_key, cfg in envs_raw.items():
        if not isinstance(cfg, dict):
            continue
        env_cfg = EnvConfig(
            name=env_key,
            label=cfg.get("label", env_key),
            external_api_base=cfg.get("external_api_base", "") or "",
            access_token=(cfg.get("access_token") or None),
            token_file=(cfg.get("token_file") or None),
            auth_audience=(cfg.get("auth_audience") or None),
            token_url=(cfg.get("token_url") or None),
            scope=(cfg.get("scope") or None),
        )
        envs[env_key] = env_cfg
        log_debug(
            f"ENV geladen: {env_key} → base={env_cfg.external_api_base!r}, "
            f"token_file={env_cfg.token_file!r}, aud={env_cfg.auth_audience!r}, "
            f"token_url={env_cfg.token_url!r}, scope={env_cfg.scope!r}"
        )

    default_env = raw.get("default_env")
    if not isinstance(default_env, str):
        default_env = None
    else:
        log_debug(f"default_env in config: {default_env}")

    if not envs:
        log_debug("Geen environments gevonden in dcbaas_api.json – dummy DEV aangemaakt.")
        envs["DEV"] = EnvConfig(
            name="DEV",
            label="DCBaaS DEV",
            external_api_base="",
            access_token=None,
            token_file=None,
            auth_audience=None,
            token_url=None,
            scope=None,
        )

    return envs, default_env


# ------------------------------------------------------------
#  Token / auth file helpers
# ------------------------------------------------------------

def load_auth_file_data(token_file: str | None) -> Dict[str, Any]:
    """
    Leest de inhoud van token_file.

    Ondersteunt:
    - JSON (met keys als 'access_token', 'jwk_path', 'client_secret', 'environment_uid', ...)
    - Plain text (enkel de token-string) → wordt in {'access_token': raw} gestopt
    """
    data: Dict[str, Any] = {}
    if not token_file:
        return data

    p = Path(token_file)
    if not p.exists():
        log_debug(f"Token file bestaat niet: {p}")
        return data

    try:
        raw = p.read_text(encoding="utf-8").strip()
    except Exception as exc:
        log_debug(f"FOUT bij lezen token file {p}: {exc}")
        return data

    if not raw:
        log_debug(f"Token file {p} is leeg.")
        return data

    # Eerst proberen als JSON
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            log_debug(
                f"Token file {p} bevat JSON met keys={list(parsed.keys())}"
            )
            return parsed
    except Exception:
        # Niet-JSON, beschouwen als plain token
        log_debug(f"Token file {p} bevat geen geldige JSON – behandeld als plain token.")
        data["access_token"] = raw
        return data

    return data


def save_auth_file_data(token_file: str | None, data: Dict[str, Any]) -> None:
    """
    Schrijft JSON terug naar token_file (als pad bestaat).
    """
    if not token_file:
        return
    p = Path(token_file)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps(data, indent=2), encoding="utf-8")
        log_debug(f"Auth data opgeslagen in {p}")
    except Exception as exc:
        log_debug(f"FOUT bij schrijven auth file {p}: {exc}")


def load_token_from_file(token_file: str | None) -> str:
    """
    Haalt 'access_token' uit token_file (JSON of plain).
    """
    data = load_auth_file_data(token_file)
    tok = data.get("access_token")
    if isinstance(tok, str) and tok.strip():
        log_debug(
            f"Access token geladen uit token_file (lengte={len(tok.strip())})"
        )
        return tok.strip()
    return ""


def load_default_token_for_env(env: EnvConfig) -> str:
    """
    Probeert een default access token op te pikken voor een omgeving:

    1) Omgevingsvariabele DCBAAS_TOKEN_<ENV>
    2) Omgevingsvariabele DCBAAS_TOKEN
    3) 'access_token' uit dcbaas_api.json (als ingevuld)
    4) 'access_token' uit token_file JSON (of plain text)
    """
    env_specific = os.getenv(f"DCBAAS_TOKEN_{env.name.upper()}")
    if env_specific:
        log_debug(
            f"Token gevonden via env var DCBAAS_TOKEN_{env.name.upper()} "
            f"(lengte={len(env_specific)})"
        )
        return env_specific.strip()

    generic = os.getenv("DCBAAS_TOKEN")
    if generic:
        log_debug(f"Token gevonden via env var DCBAAS_TOKEN (lengte={len(generic)})")
        return generic.strip()

    if env.access_token and env.access_token.strip():
        log_debug(
            f"Token gevonden in dcbaas_api.json voor {env.name} "
            f"(lengte={len(env.access_token.strip())})"
        )
        return env.access_token.strip()

    token_from_file = load_token_from_file(env.token_file)
    if token_from_file:
        log_debug(
            f"Token geladen uit token_file voor {env.name} "
            f"(lengte={len(token_from_file)})"
        )
    else:
        log_debug(f"Geen token gevonden voor {env.name} (env vars + config + file leeg).")

    return token_from_file


# ------------------------------------------------------------
#  JWT (client_assertion) genereren op basis van JWK
# ------------------------------------------------------------

def _default_audience_for_env(env_name: str) -> str:
    name_upper = env_name.upper()
    if name_upper in ("DEV", "TI"):
        return "https://authenticatie-ti.vlaanderen.be/op"
    return "https://authenticatie.vlaanderen.be/op"


def build_client_assertion_jwt(env: EnvConfig) -> Tuple[Optional[str], Optional[str]]:
    """
    Maakt een client_assertion JWT met:
      - JWK pad uit access_token.txt (key 'jwk_path')
      - iss/sub = kid uit JWK
      - aud = env.auth_audience (of fallback)
      - RS256, header.kid = kid
    """
    data = load_auth_file_data(env.token_file)
    jwk_path = data.get("jwk_path")
    if not jwk_path:
        msg = (
            f"In token_file voor omgeving {env.name} is geen 'jwk_path' gevonden. "
            "Zorg dat access_token.txt JSON bevat met minstens 'jwk_path'."
        )
        log_debug(msg)
        return None, msg

    p = Path(jwk_path)
    if not p.exists():
        msg = f"JWK-bestand niet gevonden voor omgeving {env.name}: {p}"
        log_debug(msg)
        return None, msg

    try:
        jwk_dict = json.loads(p.read_text(encoding="utf-8"))
    except Exception as exc:
        msg = f"Kon JWK JSON niet lezen voor omgeving {env.name}: {exc}"
        log_debug(msg)
        return None, msg

    kid = jwk_dict.get("kid")
    if not kid:
        msg = f"Geen 'kid' gevonden in JWK voor omgeving {env.name}."
        log_debug(msg)
        return None, msg

    aud = env.auth_audience or _default_audience_for_env(env.name)
    log_debug(f"JWT audience voor env={env.name}: {aud}")

    # kleine negatieve skew om 'iat in the future' te vermijden
    now = int(time.time()) - 10
    exp = now + 10 * 60  # 10 minuten geldig

    payload = {
        "iss": kid,
        "sub": kid,
        "iat": now,
        "exp": exp,
        "aud": aud,
    }
    log_debug(f"JWT payload voor env={env.name}: {payload}")

    try:
        jwk_json = json.dumps(jwk_dict)
        key = RSAAlgorithm.from_jwk(jwk_json)

        headers = {"typ": "JWT", "alg": "RS256", "kid": kid}
        log_debug(f"JWT headers voor env={env.name}: {headers}")

        token = jwt.encode(payload, key, algorithm="RS256", headers=headers)
        log_debug(
            f"JWT succesvol gegenereerd voor env={env.name} "
            f"(lengte={len(token) if isinstance(token, str) else 'n/a'})"
        )
        return token, None
    except Exception as exc:
        msg = f"Fout bij JWT genereren voor omgeving {env.name}: {exc}"
        log_debug(msg)
        return None, msg


# ------------------------------------------------------------
#  Access token aanvragen via /op/v1/token
# ------------------------------------------------------------

def _default_token_url_for_env(env_name: str) -> str:
    name_upper = env_name.upper()
    if name_upper in ("DEV", "TI"):
        return "https://authenticatie-ti.vlaanderen.be/op/v1/token"
    return "https://authenticatie.vlaanderen.be/op/v1/token"


def request_access_token_for_env(env: EnvConfig) -> Tuple[Optional[str], Optional[str]]:
    """
    Vraagt een nieuw access_token op bij authenticatie(-ti).vlaanderen.be
    via client_credentials + client_assertion (JWT + JWK).

    Formaat volgens documentatie:
      POST /op/v1/token
      Content-Type: application/x-www-form-urlencoded

      grant_type=client_credentials
      scope=<scopes>
      client_assertion_type=urn:ietf:params:oauth:client-assertion-type:jwt-bearer
      client_assertion=<JWT>
    """
    jwt_token, err = build_client_assertion_jwt(env)
    if err or not jwt_token:
        return None, err or "Onbekende fout bij JWT genereren."

    token_url = env.token_url or _default_token_url_for_env(env.name)
    scope = env.scope or ""
    if not scope:
        msg = (
            f"Geen 'scope' ingesteld voor omgeving {env.name} in dcbaas_api.json. "
            "Zonder scopes zal de token endpoint meestal een fout geven."
        )
        log_debug(msg)

    data = {
        "grant_type": "client_credentials",
        "scope": scope,
        "client_assertion_type": "urn:ietf:params:oauth:client-assertion-type:jwt-bearer",
        "client_assertion": jwt_token,
    }

    log_debug(
        f"POST naar token endpoint {token_url} voor env={env.name}, "
        f"scope_len={len(scope)}, jwt_len={len(jwt_token)}"
    )

    try:
        resp = requests.post(
            token_url,
            data=data,
            timeout=30,
        )
    except Exception as exc:
        msg = f"HTTP-fout bij token endpoint voor env {env.name}: {exc}"
        log_debug(msg)
        return None, msg

    log_debug(
        f"Token endpoint antwoord status={resp.status_code}, body_len={len(resp.text)}"
    )

    if resp.status_code != 200:
        short = resp.text
        if len(short) > 300:
            short = short[:300] + "..."
        msg = f"Token endpoint gaf status {resp.status_code} voor env {env.name}: {short}"
        log_debug(msg)
        return None, msg

    try:
        data_json = resp.json()
    except Exception as exc:
        msg = f"Kon JSON niet parsen van token endpoint voor env {env.name}: {exc}"
        log_debug(msg)
        return None, msg

    access_token = data_json.get("access_token")
    token_type = data_json.get("token_type", "Bearer")
    if not access_token:
        msg = (
            f"Token endpoint antwoord voor env {env.name} bevat geen 'access_token'. "
            f"JSON: {data_json}"
        )
        log_debug(msg)
        return None, msg

    full_token = f"{token_type} {access_token}".strip()
    log_debug(
        f"Nieuw access_token ontvangen voor env {env.name} "
        f"(token_type={token_type}, lengte={len(full_token)})"
    )

    # Opslaan in auth file JSON
    auth_data = load_auth_file_data(env.token_file)
    auth_data["access_token"] = full_token
    save_auth_file_data(env.token_file, auth_data)

    return full_token, None


# ------------------------------------------------------------
#  API-call naar /certificate/search
# ------------------------------------------------------------

def build_certificate_search_body(org_code: str) -> Dict[str, Any]:
    """
    Bouwt de body voor /certificate/search.
    """
    body = {
        "organization_code": org_code
    }
    log_debug(f"Request body voor org={org_code}: {body}")
    return body


def fetch_certificates_for_org(
    env: EnvConfig,
    org_code: str,
    access_token: str,
    timeout: int = 30,
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Roept /certificate/search aan voor één organisatie-code.

    access_token = exacte string die in de Authorization-header moet,
    bv. 'Bearer eyJ...'.
    """
    if not env.external_api_base:
        msg = (f"Base URL voor omgeving {env.name} is nog niet ingevuld in dcbaas_api.json. "
               f"(env.external_api_base is leeg)")
        log_debug(msg)
        return [], msg

    url = env.external_api_base.rstrip("/") + "/certificate/search"
    body = build_certificate_search_body(org_code)

    token_len = len(access_token.strip()) if access_token else 0
    log_debug(
        f"POST naar {url} voor org={org_code}, env={env.name}, "
        f"token_len={token_len}"
    )

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": access_token.strip(),
    }

    try:
        resp = requests.post(url, json=body, headers=headers, timeout=timeout)
    except Exception as exc:
        msg = f"HTTP-fout voor org {org_code} in env {env.name}: {exc}"
        log_debug(msg)
        return [], msg

    log_debug(
        f"Antwoord van {url} status={resp.status_code}, "
        f"body_len={len(resp.text)}"
    )

    if resp.status_code == 401:
        msg = (
            f"401 Unauthorized voor omgeving {env.name} (org={org_code}). "
            "Waarschijnlijk is je access token ongeldig, verlopen of ontbreekt "
            "de 'Bearer ' prefix."
        )
        log_debug(f"DETAIL 401-respons: {resp.text}")
        return [], msg

    if resp.status_code != 200:
        short = resp.text
        if len(short) > 300:
            short = short[:300] + "..."
        msg = (f"Status {resp.status_code} bij {url} voor org={org_code}: {short}")
        log_debug(msg)
        return [], msg

    try:
        data = resp.json()
    except Exception as exc:
        msg = f"Kon JSON niet parsen voor org {org_code} (env {env.name}): {exc}"
        log_debug(msg)
        return [], msg

    items = data.get("response")
    if not isinstance(items, list):
        if isinstance(data, list):
            items = data
        else:
            msg = (
                f"Onverwacht antwoord voor org {org_code} (env {env.name}): "
                f"geen 'response' lijst in JSON."
            )
            log_debug(msg + f" RAW: {data}")
            return [], msg

    log_debug(
        f"Succesvol {len(items)} certificaten ontvangen voor org {org_code} in env {env.name}."
    )
    return items, None


# ------------------------------------------------------------
#  Excel export
# ------------------------------------------------------------

def build_excel(results: Dict[str, List[Dict[str, Any]]]) -> bytes:
    """
    Maakt één XLSX met alle organisaties.
    """
    wb = Workbook()
    ws_cert = wb.active
    ws_cert.title = "Certificates"

    cert_headers = [
        "organization_code",
        "application_name",
        "application_status",
        "contact_persons",
        "description",
        "type",
        "issued_by",
        "start_date",
        "end_date",
        "status",
        "serial_number",
    ]
    ws_cert.append(cert_headers)

    total_rows = 0
    for org_code, items in results.items():
        for row in items:
            contact = row.get("contact_person") or row.get("contact_persons")
            if isinstance(contact, list):
                contact_str = ", ".join(str(c) for c in contact)
            else:
                contact_str = str(contact) if contact is not None else ""

            ws_cert.append([
                org_code,
                row.get("application_name", ""),
                row.get("application_status", ""),
                contact_str,
                row.get("description", ""),
                row.get("type", ""),
                row.get("issued_by", ""),
                row.get("start_date", ""),
                row.get("end_date", ""),
                row.get("status", ""),
                row.get("serial_number", ""),
            ])
            total_rows += 1

    log_debug(f"Excel: {total_rows} certificaat-rijen toegevoegd in 'Certificates'.")

    ws_app = wb.create_sheet("Applications")
    app_headers = [
        "organization_code",
        "application_name",
        "application_status",
        "contact_persons",
        "description",
        "type",
    ]
    ws_app.append(app_headers)

    seen = set()
    app_rows = 0
    for org_code, items in results.items():
        for row in items:
            app_name = row.get("application_name", "")
            key = (org_code, app_name)
            if key in seen:
                continue
            seen.add(key)

            contact = row.get("contact_person") or row.get("contact_persons")
            if isinstance(contact, list):
                contact_str = ", ".join(str(c) for c in contact)
            else:
                contact_str = str(contact) if contact is not None else ""

            ws_app.append([
                org_code,
                app_name,
                row.get("application_status", ""),
                contact_str,
                row.get("description", ""),
                row.get("type", ""),
            ])
            app_rows += 1

    log_debug(f"Excel: {app_rows} unieke toepassingen toegevoegd in 'Applications'.")

    for ws in (ws_cert, ws_app):
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws[f"{col_letter}{row_idx}"].value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ------------------------------------------------------------
#  Web UI
# ------------------------------------------------------------

def register_web_routes(app: Flask, settings: Dict[str, Any], tools=None) -> None:
    """
    Integreer deze tool in de bestaande CyNiT Tools Flask-app.

    Route:
      - GET/POST /dcbaas-org-export
    """
    envs, default_env = load_env_configs_from_dcbaas_api()
    log_debug(f"Environments beschikbaar: {list(envs.keys())}")

    base_css = cynit_layout.common_css(settings)
    common_js = cynit_layout.common_js()

    colors = settings.get("colors", {})
    bg = colors.get("background", "#000000")
    fg = colors.get("general_fg", "#FFFFFF")
    title_color = colors.get("title", "#00A2FF")
    t1_bg = colors.get("table_col1_bg", "#333333")
    t1_fg = colors.get("table_col1_fg", "#000000")
    t2_bg = colors.get("table_col2_bg", "#111111")
    t2_fg = colors.get("table_col2_fg", "#00FA00")
    btn_bg = colors.get("button_bg", "#111111")
    btn_fg = colors.get("button_fg", "#00B7C3")

    header = cynit_layout.header_html(
        settings,
        tools=tools,
        title="DCBaaS – Export per organisatie",
        right_html="",
    )
    footer = cynit_layout.footer_html()

    extra_css = f"""
    .card {{
      max-width: 1100px;
      margin: 0 auto 20px auto;
      background: #111111;
      padding: 20px;
      border-radius: 16px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.7);
      color: {fg};
    }}
    h1, h2 {{
      color: {title_color};
      margin-top: 0;
    }}
    label {{
      display:block;
      margin-top:12px;
      font-weight:600;
    }}
    textarea, select, input[type="text"] {{
      width:100%;
      padding:8px 10px;
      border-radius:8px;
      border:1px solid #444;
      background:{bg};
      color:{fg};
      box-sizing:border-box;
    }}
    textarea {{
      min-height:120px;
      font-family:Consolas, monospace;
    }}
    .btn {{
      margin-top:16px;
      padding:8px 16px;
      border-radius:999px;
      border:1px solid #333;
      background:{btn_bg};
      color:{btn_fg};
      font-weight:700;
      cursor:pointer;
      display:inline-block;
      margin-right:10px;
    }}
    .btn:hover {{
      filter:brightness(1.15);
    }}
    .muted {{
      color:#aaa;
      font-size:0.9em;
    }}
    .error {{
      color:#fecaca;
      background:#7f1d1d;
      padding:8px 12px;
      border-radius:8px;
      margin-bottom:10px;
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      margin-top: 10px;
      font-size:0.9em;
    }}
    th, td {{
      border: 1px solid #333;
      padding: 4px 6px;
    }}
    th {{
      background: {t1_bg};
      color: {t1_fg};
    }}
    tbody tr:nth-child(odd) {{
      background: {t2_bg};
      color: {t2_fg};
    }}
    tbody tr:nth-child(even) {{
      background: #050505;
      color: {fg};
    }}
    .jwt-box {{
      width: 100%;
      min-height: 80px;
      font-family: Consolas, monospace;
      background: {bg};
      color: {fg};
      border-radius: 8px;
      border: 1px solid #444;
      padding: 8px 10px;
      box-sizing: border-box;
      word-break: break-all;
      white-space: pre-wrap;
    }}
    """

    page_template = (
        "<!doctype html>\n"
        "<html lang='nl'>\n"
        "<head>\n"
        "  <meta charset='utf-8'>\n"
        "  <title>DCBaaS – Export per organisatie</title>\n"
        "  <style>\n"
        f"{base_css}\n{extra_css}\n"
        "  </style>\n"
        "  <script>\n"
        f"{common_js}\n"
        "  </script>\n"
        "</head>\n"
        "<body>\n"
        f"{header}\n"
        "<div class='page'>\n"
        "  <div class='card'>\n"
        "    <h1>DCBaaS – Export per organisatie</h1>\n"
        "    <p class='muted'>\n"
        "      1. Vraag (indien nodig) een nieuw access token op via JWT/JWK.<br>\n"
        "      2. Plak hieronder exact wat je ook in je andere tools als Authorization gebruikt\n"
        "         (bv. <code>Bearer eyJ...</code>).<br>\n"
        "      3. Vul één of meerdere organisatie-codes in (één per lijn) en kies Preview of Excel.<br>\n"
        "      Bij een <strong>401 Unauthorized</strong>-fout is je token waarschijnlijk ongeldig of verlopen.\n"
        "    </p>\n"
        "    {% if error %}\n"
        "      <div class='error'>{{ error }}</div>\n"
        "    {% endif %}\n"
        "    <form method='post'>\n"
        "      <label>Omgeving</label>\n"
        "      <select name='env'>\n"
        "        {% for key, env in envs.items() %}\n"
        "          <option value='{{ key }}' {% if key == current_env %}selected{% endif %}>\n"
        "            {{ key }} – {{ env.label }} ({{ env.external_api_base }})\n"
        "          </option>\n"
        "        {% endfor %}\n"
        "      </select>\n"
        "      <label>Access token (Authorization header)</label>\n"
        "      <input type='text' name='access_token' value='{{ access_token }}' />\n"
        "      <p class='muted'>Bijvoorbeeld: <code>Bearer eyJ...</code>. Laat dit niet leeg voor API-calls.</p>\n"
        "      <label>Organisatie-codes</label>\n"
        "      <textarea name='org_codes' "
        "placeholder='OVO000082&#10;OVO002949'>{{ org_input }}</textarea>\n"
        "      <p class='muted'>Lege lijnen worden genegeerd. Copy/paste uit Excel mag.</p>\n"
        "      <button type='submit' name='action' value='preview' class='btn'>Voorbeeld tonen</button>\n"
        "      <button type='submit' name='action' value='export' class='btn'>Excel downloaden</button>\n"
        "      <button type='submit' name='action' value='gen_jwt' class='btn'>Genereer client_assertion JWT</button>\n"
        "      <button type='submit' name='action' value='get_token' class='btn'>Vraag nieuw access_token op</button>\n"
        "    </form>\n"
        "  </div>\n"
        "  {% if jwt_output %}\n"
        "    <div class='card'>\n"
        "      <h2>Debug – gegenereerde client_assertion (JWT)</h2>\n"
        "      <p class='muted'>Deze JWT wordt gebruikt richting het token endpoint.</p>\n"
        "      <div class='jwt-box'>{{ jwt_output }}</div>\n"
        "    </div>\n"
        "  {% endif %}\n"
        "  {% if token_message %}\n"
        "    <div class='card'>\n"
        "      <h2>Token status</h2>\n"
        "      <p class='muted'>{{ token_message }}</p>\n"
        "    </div>\n"
        "  {% endif %}\n"
        "  {% if preview %}\n"
        "    <div class='card'>\n"
        "      <h2>Preview resultaten</h2>\n"
        "      {% if total == 0 %}\n"
        "        <p class='muted'>Geen certificaten gevonden voor de opgegeven codes.</p>\n"
        "      {% else %}\n"
        "        <p class='muted'>Totaal {{ total }} certificaten voor {{ org_count }} organisaties.</p>\n"
        "        <table>\n"
        "          <thead>\n"
        "            <tr>\n"
        "              <th>Org</th><th>Application</th><th>App status</th>\n"
        "              <th>Cert status</th><th>Serial</th><th>Start</th><th>End</th>\n"
        "            </tr>\n"
        "          </thead>\n"
        "          <tbody>\n"
        "            {% for row in preview_rows %}\n"
        "            <tr>\n"
        "              <td>{{ row.org }}</td>\n"
        "              <td>{{ row.app }}</td>\n"
        "              <td>{{ row.app_status }}</td>\n"
        "              <td>{{ row.cert_status }}</td>\n"
        "              <td>{{ row.serial }}</td>\n"
        "              <td>{{ row.start }}</td>\n"
        "              <td>{{ row.end }}</td>\n"
        "            </tr>\n"
        "            {% endfor %}\n"
        "          </tbody>\n"
        "        </table>\n"
        "      {% endif %}\n"
        "      {% if errors %}\n"
        "        <h3>Fouten / waarschuwingen</h3>\n"
        "        <ul>\n"
        "          {% for e in errors %}<li>{{ e }}</li>{% endfor %}\n"
        "        </ul>\n"
        "      {% endif %}\n"
        "    </div>\n"
        "  {% endif %}\n"
        "</div>\n"
        f"{footer}\n"
        "</body>\n"
        "</html>\n"
    )

    def _render(**ctx):
        return render_template_string(page_template, tools=tools, **ctx)

    if default_env and default_env in envs:
        initial_env = default_env
    elif "DEV" in envs:
        initial_env = "DEV"
    else:
        initial_env = next(iter(envs.keys()))
    log_debug(f"Initial environment: {initial_env}")

    @app.route("/dcbaas-org-export", methods=["GET", "POST"])
    def dcbaas_org_export():
        error: Optional[str] = None
        org_input = ""
        access_token = ""
        preview = False
        preview_rows: List[Dict[str, Any]] = []
        errors: List[str] = []
        jwt_output = ""
        token_message = ""
        total = 0

        current_env_key = initial_env

        if request.method == "GET":
            env = envs.get(current_env_key, next(iter(envs.values())))
            access_token_local = load_default_token_for_env(env)
            log_debug(
                f"GET /dcbaas-org-export voor env={env.name}, "
                f"default_token_len={len(access_token_local) if access_token_local else 0}"
            )
            return _render(
                envs=envs,
                current_env=current_env_key,
                error=None,
                org_input="",
                access_token=access_token_local,
                preview=False,
                preview_rows=[],
                total=0,
                org_count=0,
                errors=[],
                jwt_output="",
                token_message="",
            )

        # POST
        current_env_key = request.form.get("env", current_env_key)
        env = envs.get(current_env_key, next(iter(envs.values())))

        org_input = request.form.get("org_codes", "") or ""
        access_token = request.form.get("access_token", "") or ""
        action = request.form.get("action") or "preview"

        org_codes = [line.strip() for line in org_input.splitlines() if line.strip()]

        log_debug(
            f"POST /dcbaas-org-export action={action}, env={env.name}, "
            f"#orgs={len(org_codes)}, token_len={len(access_token.strip()) if access_token else 0}"
        )

        results_by_org: Dict[str, List[Dict[str, Any]]] = {}

        if action == "gen_jwt":
            jwt_token, err = build_client_assertion_jwt(env)
            if err:
                errors.append(err)
            else:
                jwt_output = jwt_token
                token_message = (
                    "client_assertion JWT succesvol gegenereerd. "
                    "Gebruik deze in je token-request of voor debug."
                )

        elif action == "get_token":
            new_token, err = request_access_token_for_env(env)
            if err:
                errors.append(err)
            else:
                access_token = new_token
                token_message = (
                    f"Nieuw access_token succesvol opgehaald en opgeslagen in token_file "
                    f"voor omgeving {env.name}."
                )

        else:
            # Validatie voor echte API-calls
            if not access_token.strip():
                error = "Geef een access token in (Authorization header waarde)."
            elif not org_codes:
                error = "Geef minstens één organisatie-code in."

            if not error:
                for org in org_codes:
                    items, err = fetch_certificates_for_org(env, org, access_token)
                    if err:
                        errors.append(err)
                    results_by_org[org] = items
                    total += len(items)

                if action == "export":
                    log_debug(f"Excel-export gevraagd voor env={env.name}, totaal={total} certificaten.")
                    xlsx_bytes = build_excel(results_by_org)
                    buf = BytesIO(xlsx_bytes)
                    buf.seek(0)
                    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
                    filename = f"dcbaas_org_export_{current_env_key}_{ts}.xlsx"
                    return send_file(
                        buf,
                        as_attachment=True,
                        download_name=filename,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # Preview
                preview = True
                max_rows = 50
                for org, items in results_by_org.items():
                    for row in items:
                        preview_rows.append({
                            "org": org,
                            "app": row.get("application_name", ""),
                            "app_status": row.get("application_status", ""),
                            "cert_status": row.get("status", ""),
                            "serial": row.get("serial_number", ""),
                            "start": row.get("start_date", ""),
                            "end": row.get("end_date", ""),
                        })
                        if len(preview_rows) >= max_rows:
                            break
                    if len(preview_rows) >= max_rows:
                        break

                log_debug(
                    f"Preview: {len(preview_rows)} rijen getoond (totaal={total}) "
                    f"voor {len(org_codes)} organisaties."
                )

        org_count = len({r["org"] for r in preview_rows}) if preview_rows else 0

        return _render(
            envs=envs,
            current_env=current_env_key,
            error=error,
            org_input=org_input,
            access_token=access_token,
            preview=preview,
            preview_rows=preview_rows,
            total=total,
            org_count=org_count,
            errors=errors,
            jwt_output=jwt_output,
            token_message=token_message,
        )


# Standalone web-run (optioneel)
if __name__ == "__main__":
    settings = cynit_theme.load_settings()
    tools_cfg = cynit_theme.load_tools()
    tools = tools_cfg.get("tools", [])
    app = Flask(__name__)
    register_web_routes(app, settings, tools)
    app.run(host="127.0.0.1", port=5451, debug=True)
